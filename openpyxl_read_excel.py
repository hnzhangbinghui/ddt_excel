import openpyxl

wb = openpyxl.load_workbook('test_excel.xlsx')
print(type(wb))
sheets=wb.sheetnames
print(sheets)
sheet=wb['Sheet1']
print(sheet)
print(sheet['A1'].value)
print(sheet.max_row)
print(sheet.max_column)
# print(sheet.cell(1,2).value)
l=[]
for i in range(2,sheet.max_row+1):
    # print(sheet.cell(i,2).value)
    l.append(sheet.cell(i,2).value)
print(l)

# np.savetxt(r'C:\Users\LX\Desktop\test_csv.csv',a,fmt='%d',delimiter=',',header="A,B,C,D,E,F,G,H,I,L,J,K,L,M,N,O,P,Q,R,S")

import numpy as np
aa=np.savetxt('zhang.txt',l,header='passwd',fmt='%d')
print(aa)

